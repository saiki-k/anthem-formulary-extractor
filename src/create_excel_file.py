import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_excel_from_json(json_path, output_path):
    """
    Create an Excel file with multiple sheets from extracted JSON data.

    Args:
        json_path: Path to the extracted_data.json file
        output_path: Path where the Excel file should be saved
    """
    # Load the JSON data
    with open(json_path, "r", encoding="utf-8") as f:
        categories_data = json.load(f)

    # Create a new workbook
    wb = openpyxl.Workbook()

    # Remove the default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create the Categories index sheet
    categories_sheet = wb.create_sheet("Categories", 0)

    # Style for the Categories sheet header
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)

    # Add header to Categories sheet
    categories_sheet["A1"] = "Category"
    categories_sheet["B1"] = "Total Subcategories"
    categories_sheet["C1"] = "Total Drugs"
    categories_sheet["D1"] = "Link"

    for cell in ["A1", "B1", "C1", "D1"]:
        categories_sheet[cell].fill = header_fill
        categories_sheet[cell].font = header_font
        categories_sheet[cell].alignment = Alignment(
            horizontal="center", vertical="center"
        )

    # Set column widths for Categories sheet
    categories_sheet.column_dimensions["A"].width = 60
    categories_sheet.column_dimensions["B"].width = 20
    categories_sheet.column_dimensions["C"].width = 15
    categories_sheet.column_dimensions["D"].width = 20

    # Freeze the header row so it stays visible when scrolling
    categories_sheet.freeze_panes = "A2"

    print(f"\nCreating Excel file with {len(categories_data)} categories...")

    # Process each category
    for idx, category in enumerate(categories_data, start=2):
        category_name = category["categoryName"]
        subcategories = category["subCategories"]

        # Count total drugs in this category
        total_drugs = sum(len(subcat["rows"]) for subcat in subcategories)

        # Create sanitized sheet name (Excel has 31 char limit and invalid chars)
        sheet_name = category_name[:31] if len(category_name) > 31 else category_name
        # Replace invalid characters for sheet names
        for char in ["\\", "/", "*", "[", "]", ":", "?"]:
            sheet_name = sheet_name.replace(char, "-")

        # Add category name (no link)
        categories_sheet[f"A{idx}"] = category_name
        categories_sheet[f"A{idx}"].alignment = Alignment(
            horizontal="left", vertical="center"
        )

        # Add subcategory count
        categories_sheet[f"B{idx}"] = len(subcategories)
        categories_sheet[f"B{idx}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # Add total drugs count
        categories_sheet[f"C{idx}"] = total_drugs
        categories_sheet[f"C{idx}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # Add clickable link in separate column
        link_cell = categories_sheet[f"D{idx}"]
        link_cell.value = "Go to sheet"
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = Font(color="0563C1", underline="single")
        link_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Create a sheet for this category
        cat_sheet = wb.create_sheet(sheet_name)

        # Add header row
        cat_sheet["A1"] = "Category"
        cat_sheet["B1"] = "Subcategory"
        cat_sheet["C1"] = "Drug"
        cat_sheet["D1"] = "Tier"
        cat_sheet["E1"] = "Notes"

        for cell in ["A1", "B1", "C1", "D1", "E1"]:
            cat_sheet[cell].fill = header_fill
            cat_sheet[cell].font = header_font
            cat_sheet[cell].alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Set column widths to fit content
        cat_sheet.column_dimensions["A"].width = 45  # Category
        cat_sheet.column_dimensions["B"].width = 45  # Subcategory
        cat_sheet.column_dimensions["C"].width = 65  # Drug
        cat_sheet.column_dimensions["D"].width = 15  # Tier
        cat_sheet.column_dimensions["E"].width = 35  # Notes

        # Freeze the header row so it stays visible when scrolling
        cat_sheet.freeze_panes = "A2"

        # Add data rows
        current_row = 2
        for subcat in subcategories:
            subcat_name = subcat["subCategoryName"]
            rows = subcat["rows"]

            for row_data in rows:
                cat_sheet[f"A{current_row}"] = category_name
                cat_sheet[f"B{current_row}"] = subcat_name
                cat_sheet[f"C{current_row}"] = row_data["drug_name"]
                cat_sheet[f"D{current_row}"] = row_data["tier"]
                cat_sheet[f"E{current_row}"] = row_data["notes"]

                # Alignment
                cat_sheet[f"A{current_row}"].alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
                cat_sheet[f"B{current_row}"].alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
                cat_sheet[f"C{current_row}"].alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
                cat_sheet[f"D{current_row}"].alignment = Alignment(
                    horizontal="center", vertical="top"
                )
                cat_sheet[f"E{current_row}"].alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

                current_row += 1

        print(
            f"  Created sheet: {sheet_name} ({len(subcategories)} subcategories, {total_drugs} drugs)"
        )

    # Save the workbook
    wb.save(output_path)
    print(f"\nExcel file saved to: {output_path}")
    print(f"  Total sheets created: {len(wb.sheetnames)}")

    return output_path


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Create Excel file from extracted PDF data."
    )
    parser.add_argument(
        "json_path",
        help="Path to the extracted_data.json file",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Output path for Excel file (default: same directory as JSON with .xlsx extension)",
    )

    args = parser.parse_args()

    json_path = Path(args.json_path)

    if not json_path.exists():
        print(f"Error: JSON file not found: {json_path}")
        return 1

    # Determine output path
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = json_path.parent / f"{json_path.parent.name}.xlsx"

    create_excel_from_json(json_path, output_path)

    return 0


if __name__ == "__main__":
    exit(main())
